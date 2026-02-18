#!/usr/bin/env python3
"""
Fase R6 STAP A: Excel Generator voor Frank's Handmatige Review
HolidaiButler Content Repair Pipeline

Genereert een werkbaar Excel-bestand met Top 150 meest zichtbare POIs
uit de pending + review_required staging entries.

Usage:
    python3 fase_r6_excel_generator.py
"""

import json
import sys
import os
from datetime import datetime

import mysql.connector

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# === DATABASE CONFIG ===
DB_CONFIG = {
    'host': 'jotx.your-database.de',
    'user': 'pxoziy_1',
    'password': 'j8,DrtshJSm$',
    'database': 'pxoziy_db1',
    'charset': 'utf8mb4',
    'collation': 'utf8mb4_unicode_ci',
}

# R4 batch date
R4_BATCH_DATE = '2026-02-13'

# Output paths
OUTPUT_XLSX = '/root/fase_r6_frank_review.xlsx'
OUTPUT_JSON = '/root/fase_r6_top150_raw.json'

# Maximum POIs to include
MAX_POIS = 150


def log(msg):
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", flush=True)


def get_connection():
    return mysql.connector.connect(**DB_CONFIG)


def fetch_top_pois(conn, limit=MAX_POIS):
    """Fetch top POIs by visibility score from pending + review_required staging."""
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            p.id as poi_id,
            p.name,
            p.category,
            p.subcategory,
            p.poi_type,
            CASE p.destination_id WHEN 1 THEN 'Calpe' WHEN 2 THEN 'Texel' END as bestemming,
            p.rating,
            p.review_count,
            ROUND((COALESCE(p.rating, 0)/5 * 0.4) + (LOG10(COALESCE(p.review_count, 0)+1)/3 * 0.6), 4) as visibility_score,
            p.website,
            p.address,
            s.status as staging_status,
            s.detail_description_en as nieuwe_tekst,
            s.comparison_recommendation,
            p.enriched_detail_description as huidige_productie_tekst,
            s.id as staging_id,
            p.destination_id
        FROM poi_content_staging s
        JOIN POI p ON s.poi_id = p.id
        WHERE s.status IN ('pending', 'review_required')
          AND DATE(s.created_at) = %s
          AND p.is_active = 1
        ORDER BY
            ROUND((COALESCE(p.rating, 0)/5 * 0.4) + (LOG10(COALESCE(p.review_count, 0)+1)/3 * 0.6), 4) DESC
        LIMIT %s
    """, (R4_BATCH_DATE, limit))

    rows = cursor.fetchall()
    cursor.close()
    return rows


def create_excel(rows, output_path):
    """Create the formatted Excel workbook."""
    wb = Workbook()

    # =====================================================
    # SHEET 1: POI Review
    # =====================================================
    ws = wb.active
    ws.title = "POI Review"

    # --- Styles ---
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(vertical='top', wrap_text=True)

    readonly_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    editable_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    # Conditional formatting colors for column K
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    orange_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Texel/Calpe border colors
    texel_border = Border(left=Side(style='thick', color='30C59B'))
    calpe_border = Border(left=Side(style='thick', color='7FA594'))

    # Hyperlink font
    link_font = Font(name='Arial', size=10, color='0563C1', underline='single')

    # --- Column definitions ---
    # Note: facebook/instagram not in POI table, using website only
    columns = [
        ('A', 'poi_id', 10),
        ('B', 'staging_id', 10),
        ('C', 'Bestemming', 12),
        ('D', 'Naam', 30),
        ('E', 'Categorie', 20),
        ('F', 'Subcategorie', 20),
        ('G', 'Rating', 8),
        ('H', 'Reviews', 8),
        ('I', 'Website', 30),
        ('J', 'Nieuwe Tekst (EN)', 60),
        ('K', 'BEOORDELING', 15),
        ('L', 'AANGEPASTE TEKST', 60),
        ('M', 'OPMERKINGEN', 30),
        ('N', 'Huidige Tekst', 60),
        ('O', 'Staging Status', 12),
    ]

    # --- Write headers ---
    for col_idx, (col_letter, header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[col_letter].width = width

    # --- Freeze panes (row 1) ---
    ws.freeze_panes = 'A2'

    # --- Auto filter ---
    ws.auto_filter.ref = f"A1:O{len(rows) + 1}"

    # --- Data validation for column K (BEOORDELING) ---
    dv = DataValidation(
        type='list',
        formula1='"GOED,AANPASSEN,AFKEUREN"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = 'Kies GOED, AANPASSEN of AFKEUREN'
    dv.errorTitle = 'Ongeldige keuze'
    dv.prompt = 'Kies: GOED, AANPASSEN of AFKEUREN'
    dv.promptTitle = 'Beoordeling'
    ws.add_data_validation(dv)

    # --- Write data rows ---
    for row_idx, poi in enumerate(rows, 2):
        # Column A: poi_id
        ws.cell(row=row_idx, column=1, value=poi['poi_id']).font = data_font

        # Column B: staging_id
        ws.cell(row=row_idx, column=2, value=poi['staging_id']).font = data_font

        # Column C: Bestemming
        ws.cell(row=row_idx, column=3, value=poi['bestemming']).font = data_font

        # Column D: Naam
        ws.cell(row=row_idx, column=4, value=poi['name']).font = data_font

        # Column E: Categorie
        ws.cell(row=row_idx, column=5, value=poi['category']).font = data_font

        # Column F: Subcategorie
        ws.cell(row=row_idx, column=6, value=poi['subcategory']).font = data_font

        # Column G: Rating
        cell_g = ws.cell(row=row_idx, column=7, value=poi['rating'])
        cell_g.font = data_font
        cell_g.number_format = '0.0'

        # Column H: Reviews
        ws.cell(row=row_idx, column=8, value=poi['review_count']).font = data_font

        # Column I: Website (hyperlink)
        website = poi.get('website') or ''
        cell_i = ws.cell(row=row_idx, column=9, value=website)
        if website and website.startswith('http'):
            cell_i.hyperlink = website
            cell_i.font = link_font
        else:
            cell_i.font = data_font

        # Column J: Nieuwe Tekst (EN)
        ws.cell(row=row_idx, column=10, value=poi['nieuwe_tekst']).font = data_font

        # Column K: BEOORDELING (editable, data validation)
        cell_k = ws.cell(row=row_idx, column=11, value='')
        cell_k.font = Font(name='Arial', size=11, bold=True)
        dv.add(cell_k)

        # Column L: AANGEPASTE TEKST (editable)
        ws.cell(row=row_idx, column=12, value='').font = data_font

        # Column M: OPMERKINGEN (editable)
        ws.cell(row=row_idx, column=13, value='').font = data_font

        # Column N: Huidige Tekst
        ws.cell(row=row_idx, column=14, value=poi['huidige_productie_tekst']).font = data_font

        # Column O: Staging Status
        ws.cell(row=row_idx, column=15, value=poi['staging_status']).font = data_font

        # --- Apply fills and borders ---
        # Read-only columns: A-J, N-O
        for col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 14, 15]:
            ws.cell(row=row_idx, column=col).fill = readonly_fill
            ws.cell(row=row_idx, column=col).alignment = data_alignment

        # Editable columns: K, L, M
        for col in [11, 12, 13]:
            ws.cell(row=row_idx, column=col).fill = editable_fill
            ws.cell(row=row_idx, column=col).alignment = data_alignment

        # Destination border
        border = texel_border if poi['bestemming'] == 'Texel' else calpe_border
        ws.cell(row=row_idx, column=1).border = border

    # --- Row height for text columns ---
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 60

    # --- Conditional formatting on column K ---
    last_row = len(rows) + 1
    k_range = f"K2:K{last_row}"

    ws.conditional_formatting.add(k_range, CellIsRule(
        operator='equal', formula=['"GOED"'], fill=green_fill
    ))
    ws.conditional_formatting.add(k_range, CellIsRule(
        operator='equal', formula=['"AANPASSEN"'], fill=orange_fill
    ))
    ws.conditional_formatting.add(k_range, CellIsRule(
        operator='equal', formula=['"AFKEUREN"'], fill=red_fill
    ))

    # --- Hide columns B and O ---
    ws.column_dimensions['B'].hidden = True
    ws.column_dimensions['O'].hidden = True

    # =====================================================
    # SHEET 2: Instructies
    # =====================================================
    ws2 = wb.create_sheet("Instructies")

    instructions = [
        ("INSTRUCTIES VOOR BEOORDELING", True, 16),
        ("", False, 11),
        (f"Dit Excel-bestand bevat de Top {len(rows)} meest zichtbare POIs", False, 11),
        ("van Texel en Calpe die nog niet in productie staan.", False, 11),
        ("", False, 11),
        ("HOE TE BEOORDELEN:", True, 13),
        ("", False, 11),
        ("1. Lees de \"Nieuwe Tekst (EN)\" in kolom J", False, 11),
        ("2. Vergelijk eventueel met de website (kolom I - klikbaar)", False, 11),
        ("3. Kies in kolom K (gele kolom) een van drie opties:", False, 11),
        ("", False, 11),
        ("   GOED        - De tekst is correct en mag naar productie", False, 11),
        ("   AANPASSEN   - De tekst bevat fouten. Typ je verbeterde", False, 11),
        ("                  versie in kolom L (gele kolom)", False, 11),
        ("   AFKEUREN    - De tekst is onbruikbaar. Er wordt een", False, 11),
        ("                  generieke veilige beschrijving gebruikt.", False, 11),
        ("", False, 11),
        ("4. Kolom M (Opmerkingen) is optioneel - voor je eigen notities", False, 11),
        ("", False, 11),
        ("BELANGRIJK:", True, 13),
        ("", False, 11),
        ("  - Beoordeel ALLEEN kolom K voor elke rij", False, 11),
        ("  - Bij \"AANPASSEN\": typ de VOLLEDIGE nieuwe tekst in kolom L", False, 11),
        ("  - Laat GEEN rijen onbeoordeeld - elke POI heeft een keuze nodig", False, 11),
        ("  - Sla het bestand op als .xlsx (niet .csv)", False, 11),
        ("  - Upload het beoordeelde bestand terug naar Claude", False, 11),
        ("", False, 11),
        ("Na upload verwerkt Claude automatisch:", True, 12),
        ("  - GOED -> content naar productie", False, 11),
        ("  - AANPASSEN -> jouw tekst naar productie", False, 11),
        ("  - AFKEUREN -> generieke veilige beschrijving", False, 11),
        ("", False, 11),
        ("Daarna worden alle teksten automatisch vertaald naar NL, DE en ES.", False, 11),
    ]

    ws2.column_dimensions['A'].width = 80

    for row_idx, (text, is_bold, size) in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        cell.font = Font(name='Arial', size=size, bold=is_bold)
        cell.alignment = Alignment(wrap_text=True)

    # =====================================================
    # SHEET 3: Statistieken
    # =====================================================
    ws3 = wb.create_sheet("Statistieken")

    # Count per destination
    texel_count = sum(1 for r in rows if r['bestemming'] == 'Texel')
    calpe_count = sum(1 for r in rows if r['bestemming'] == 'Calpe')

    # Count per category
    categories = {}
    for r in rows:
        cat = r['category'] or 'Onbekend'
        categories[cat] = categories.get(cat, 0) + 1

    # Header style
    stat_header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    stat_header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    stat_font = Font(name='Arial', size=11)
    stat_bold = Font(name='Arial', size=11, bold=True)

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 25

    # Dashboard header
    cell = ws3.cell(row=1, column=1, value="Metriek")
    cell.font = stat_header_font
    cell.fill = stat_header_fill
    cell = ws3.cell(row=1, column=2, value="Waarde")
    cell.font = stat_header_font
    cell.fill = stat_header_fill

    stats = [
        ("Totaal POIs in dit bestand", len(rows)),
        ("Waarvan Texel", texel_count),
        ("Waarvan Calpe", calpe_count),
        ("", ""),
        ("Nog te beoordelen", f'=COUNTBLANK(\'POI Review\'!K2:K{last_row})'),
        ("Beoordeeld als GOED", f'=COUNTIF(\'POI Review\'!K2:K{last_row},"GOED")'),
        ("Beoordeeld als AANPASSEN", f'=COUNTIF(\'POI Review\'!K2:K{last_row},"AANPASSEN")'),
        ("Beoordeeld als AFKEUREN", f'=COUNTIF(\'POI Review\'!K2:K{last_row},"AFKEUREN")'),
    ]

    for row_idx, (label, value) in enumerate(stats, 2):
        cell_a = ws3.cell(row=row_idx, column=1, value=label)
        cell_a.font = stat_bold if label else stat_font

        cell_b = ws3.cell(row=row_idx, column=2, value=value)
        cell_b.font = stat_font
        cell_b.alignment = Alignment(horizontal='center')

    # Category breakdown
    cat_start = len(stats) + 3
    cell = ws3.cell(row=cat_start, column=1, value="Categorie")
    cell.font = stat_header_font
    cell.fill = stat_header_fill
    cell = ws3.cell(row=cat_start, column=2, value="Aantal")
    cell.font = stat_header_font
    cell.fill = stat_header_fill

    for idx, (cat, count) in enumerate(sorted(categories.items(), key=lambda x: -x[1]), cat_start + 1):
        ws3.cell(row=idx, column=1, value=cat).font = stat_font
        cell_b = ws3.cell(row=idx, column=2, value=count)
        cell_b.font = stat_font
        cell_b.alignment = Alignment(horizontal='center')

    # --- Save ---
    wb.save(output_path)
    log(f"Excel opgeslagen: {output_path}")
    return len(rows), texel_count, calpe_count, categories


def main():
    log("=" * 70)
    log("FASE R6 STAP A: EXCEL GENERATOR VOOR FRANK'S REVIEW")
    log("HolidaiButler Content Repair Pipeline")
    log("=" * 70)

    conn = get_connection()

    try:
        # Step 1: Fetch Top 150 POIs
        log(f"\nStap A.1: Top {MAX_POIS} POIs ophalen...")
        rows = fetch_top_pois(conn, MAX_POIS)
        log(f"  Gevonden: {len(rows)} POIs")

        if not rows:
            log("ERROR: Geen POIs gevonden in staging!")
            return

        # Save raw JSON
        log(f"\nRaw data opslaan: {OUTPUT_JSON}")
        json_rows = []
        for r in rows:
            row_copy = dict(r)
            # Convert Decimal to float for JSON
            for k, v in row_copy.items():
                if hasattr(v, 'as_integer_ratio'):  # Decimal/float
                    row_copy[k] = float(v)
            json_rows.append(row_copy)

        with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
            json.dump(json_rows, f, ensure_ascii=False, indent=2, default=str)
        log(f"  Opgeslagen: {len(json_rows)} records")

        # Step 2: Generate Excel
        log(f"\nStap A.2: Excel genereren...")
        total, texel, calpe, categories = create_excel(rows, OUTPUT_XLSX)

        # Step 3: Report
        log("\n" + "=" * 70)
        log("RAPPORT VOOR FRANK")
        log("=" * 70)
        log(f"\nExcel-bestand: {OUTPUT_XLSX}")
        log(f"Totaal POIs: {total}")
        log(f"  Texel: {texel}")
        log(f"  Calpe: {calpe}")
        log(f"\nVerdeling per categorie:")
        for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
            log(f"  {cat:35s}: {count}")

        avg_time_min = total * 1.5  # 1-2 min per POI, average 1.5
        hours = int(avg_time_min // 60)
        mins = int(avg_time_min % 60)
        log(f"\nGeschatte reviewtijd: {hours} uur en {mins} minuten")
        log(f"  (gebaseerd op ~1-2 minuten per POI)")

        # Visibility score stats
        scores = [float(r['visibility_score']) for r in rows if r['visibility_score']]
        if scores:
            log(f"\nVisibility scores:")
            log(f"  Hoogste: {max(scores):.4f}")
            log(f"  Laagste: {min(scores):.4f}")
            log(f"  Gemiddeld: {sum(scores)/len(scores):.4f}")

        # Status breakdown
        pending = sum(1 for r in rows if r['staging_status'] == 'pending')
        review_req = sum(1 for r in rows if r['staging_status'] == 'review_required')
        log(f"\nStaging status verdeling:")
        log(f"  pending: {pending}")
        log(f"  review_required: {review_req}")

    finally:
        conn.close()

    log("\n" + "=" * 70)
    log("STAP A AFGEROND")
    log("=" * 70)
    log("\n>>> PAUZEER — Wacht op Frank's beoordeeld Excel-bestand <<<")


if __name__ == '__main__':
    main()
