#!/usr/bin/env python3
"""
Fase R6b STAP 4: Frank's Steekproef Excel (20 POIs)
====================================================
Genereer Excel-bestand met 20 willekeurige POIs uit de 2.047
gestrippte POIs voor Frank's handmatige controle.
10 Texel + 10 Calpe, verspreid over categorieën.
"""

import json
import sys
import argparse
from datetime import datetime

import mysql.connector

DB_CONFIG = {
    'host': 'jotx.your-database.de',
    'user': 'pxoziy_1',
    'password': 'j8,DrtshJSm$',
    'database': 'pxoziy_db1',
    'charset': 'utf8mb4',
    'collation': 'utf8mb4_unicode_ci'
}


def log(msg):
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", flush=True)


def main():
    parser = argparse.ArgumentParser(description='Fase R6b STAP 4: Steekproef Excel')
    parser.add_argument('--output', type=str, default='/root/fase_r6b_steekproef.xlsx',
                        help='Output Excel pad')
    args = parser.parse_args()

    log("=" * 70)
    log("FASE R6b STAP 4: FRANK'S STEEKPROEF EXCEL")
    log("=" * 70)

    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)

    # Selecteer 10 Texel + 10 Calpe, verspreid over categorieën, hoge visibility
    # Gebruik de audit trail om gestrippte POIs te identificeren
    cursor.execute("""
        (SELECT p.id, p.name, p.category, 'Texel' as bestemming,
                p.rating, p.review_count, p.website,
                p.enriched_detail_description as nieuwe_tekst_en,
                p.enriched_detail_description_nl as nieuwe_tekst_nl,
                h.old_value as oude_tekst_en
         FROM POI p
         JOIN poi_content_history h ON p.id = h.poi_id
         WHERE h.change_source = 'r6b_claim_strip'
           AND h.field_name = 'enriched_detail_description'
           AND p.destination_id = 2
         ORDER BY p.review_count DESC
         LIMIT 10)
        UNION ALL
        (SELECT p.id, p.name, p.category, 'Calpe' as bestemming,
                p.rating, p.review_count, p.website,
                p.enriched_detail_description as nieuwe_tekst_en,
                p.enriched_detail_description_nl as nieuwe_tekst_nl,
                h.old_value as oude_tekst_en
         FROM POI p
         JOIN poi_content_history h ON p.id = h.poi_id
         WHERE h.change_source = 'r6b_claim_strip'
           AND h.field_name = 'enriched_detail_description'
           AND p.destination_id = 1
         ORDER BY p.review_count DESC
         LIMIT 10)
    """)
    sample = cursor.fetchall()
    log(f"Steekproef: {len(sample)} POIs geselecteerd")

    if len(sample) == 0:
        log("ERROR: Geen gestrippte POIs gevonden in audit trail.")
        log("Zorg dat STAP 2 (claim stripping) en --apply-db zijn uitgevoerd.")
        cursor.close()
        conn.close()
        return

    # Genereer Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        log("openpyxl niet geïnstalleerd. Installeren...")
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "R6b Steekproef"

    # Headers
    headers = [
        ('A', 'POI ID', 10),
        ('B', 'Bestemming', 12),
        ('C', 'Naam', 30),
        ('D', 'Categorie', 20),
        ('E', 'Rating', 8),
        ('F', 'Reviews', 8),
        ('G', 'Website', 35),
        ('H', 'OUDE Tekst (EN)', 55),
        ('I', 'NIEUWE Tekst (EN)', 55),
        ('J', 'OORDEEL', 12),
        ('K', 'OPMERKING', 40),
    ]

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    review_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_letter, header_text, width in headers:
        cell = ws[f"{col_letter}1"]
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[col_letter].width = width

    # Data
    for row_idx, poi in enumerate(sample, start=2):
        ws[f"A{row_idx}"] = poi['id']
        ws[f"B{row_idx}"] = poi['bestemming']
        ws[f"C{row_idx}"] = poi['name']
        ws[f"D{row_idx}"] = poi['category']
        ws[f"E{row_idx}"] = poi['rating']
        ws[f"F{row_idx}"] = poi['review_count']
        ws[f"G{row_idx}"] = poi['website'] or ''
        ws[f"H{row_idx}"] = poi['oude_tekst_en'] or ''
        ws[f"I{row_idx}"] = poi['nieuwe_tekst_en'] or ''
        ws[f"J{row_idx}"] = ''
        ws[f"K{row_idx}"] = ''

        # Styling
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            cell = ws[f"{col}{row_idx}"]
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col in ['H', 'I', 'K']))

        # Gele achtergrond voor review kolommen
        ws[f"J{row_idx}"].fill = review_fill
        ws[f"K{row_idx}"].fill = review_fill

    # Dropdown voor OORDEEL kolom
    dv = DataValidation(
        type="list",
        formula1='"GOED,AANPASSEN,AFKEUREN"',
        allow_blank=True
    )
    dv.error = "Kies GOED, AANPASSEN of AFKEUREN"
    dv.errorTitle = "Ongeldig oordeel"
    ws.add_data_validation(dv)
    dv.add(f"J2:J{len(sample) + 1}")

    # Freeze panes
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:K{len(sample) + 1}"

    # Opslaan
    wb.save(args.output)
    log(f"Excel opgeslagen: {args.output}")

    # Statistieken
    log(f"\nSteekproef overzicht:")
    texel = [p for p in sample if p['bestemming'] == 'Texel']
    calpe = [p for p in sample if p['bestemming'] == 'Calpe']
    log(f"  Texel: {len(texel)} POIs")
    log(f"  Calpe: {len(calpe)} POIs")

    categories = {}
    for p in sample:
        cat = p['category'] or 'Unknown'
        categories[cat] = categories.get(cat, 0) + 1
    log(f"  Categorieën: {categories}")

    cursor.close()
    conn.close()


if __name__ == '__main__':
    main()
